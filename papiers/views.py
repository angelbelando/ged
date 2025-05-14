
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import ListView, DetailView, TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.db.models import Q
import fitz  # PyMuPDF
from PIL import Image
from io import BytesIO
from .models import Document
import openpyxl
from openpyxl.utils import get_column_letter

def generate_thumbnail(request, document_id):
    document = get_object_or_404(Document, id=document_id)
    doc = fitz.open(document.Document_pdf.path)
    page = doc.load_page(0)  # Charger la première page
    pix = page.get_pixmap()
    img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
    img.thumbnail((200, 200))  # Redimensionner l'image
    thumbnail_io = BytesIO()
    img.save(thumbnail_io, format='JPEG')
    thumbnail_io.seek(0)
    return HttpResponse(thumbnail_io, content_type='image/jpeg')

def export_to_excel(request):
    # Créer un classeur Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Documents"

    # Ajouter les en-têtes de colonnes
    columns = ['Rubrique', 'Type_document', 'Document', 'Date de référence', 'commentaire', 'Conservation']
    for col_num, column_title in enumerate(columns, 1):
        col_letter = get_column_letter(col_num)
        sheet[f"{col_letter}1"] = column_title

    # Ajouter les données
    documents = Document.objects.all().order_by('rubrique','document')
    for row_num, document in enumerate(documents, start=2):
        sheet[f"A{row_num}"] = document.rubrique
        sheet[f"B{row_num}"] = document.type_document.name if document.type_document else '' 
        sheet[f"C{row_num}"] = document.document
        sheet[f"D{row_num}"] = document.date_reference
        sheet[f"E{row_num}"] = document.commentaire
        sheet[f"F{row_num}"] = document.est_conserve()
     
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename="documents.xlsx"'

    workbook.save(response)
    return response

class DocumentListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = Document
    context_object_name = "documents"
    template_name = 'documents/liste_documents.html'
    permission_required = 'papiers.view_document'
    permission_denied_message = "Vous n'avez pas la permission de voir cette page."
    paginate_by = 10
    
    def get_queryset(self):
        query = self.request.GET.get('q')
        if query:
            return Document.objects.filter(
                Q(document__icontains=query) | Q(rubrique__icontains=query)
                | Q(commentaire__icontains=query)
            )
        return Document.objects.all().order_by('rubrique','document')
    
class DocumenttDetailView(DetailView):
    model = Document
    template_name = 'documents/detail_document.html'
