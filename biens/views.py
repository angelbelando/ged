from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django_filters.views import FilterView
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from .models import Objet, Categorie, Rubrique
from django.views.generic import ListView, DetailView, TemplateView
import fitz  # PyMuPDF
from PIL import Image
from io import BytesIO
from django.db.models import Sum
import matplotlib.pyplot as plt
import base64
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.db.models import Q
import openpyxl
from openpyxl.utils import get_column_letter
# from openpyxl.drawing.image import Image

def generate_thumbnail(request, objet_id):
    objet = get_object_or_404(Objet, id=objet_id)
    doc = fitz.open(objet.document.path)
    page = doc.load_page(0)  # Charger la première page
    pix = page.get_pixmap()
    img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
    img.thumbnail((200, 200))  # Redimensionner l'image
    thumbnail_io = BytesIO()
    img.save(thumbnail_io, format='JPEG')
    thumbnail_io.seek(0)
    return HttpResponse(thumbnail_io, content_type='image/jpeg')

def export_to_excel(request):
    # Créer un classeur Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Objets"

    # Ajouter les en-têtes de colonnes
    columns = ['Pièce', 'ID', 'Nom', 'Rubrique', 'Catégorie','Prix']
    for col_num, column_title in enumerate(columns, 1):
        col_letter = get_column_letter(col_num)
        sheet[f"{col_letter}1"] = column_title

    # Ajouter les données
    objets = Objet.objects.all().order_by('piece','rubrique__name','name')
    for row_num, objet in enumerate(objets, start=2):
        sheet[f"A{row_num}"] = objet.piece
        sheet[f"B{row_num}"] = objet.id
        sheet[f"C{row_num}"] = objet.name
        sheet[f"D{row_num}"] = objet.rubrique.name if objet.rubrique else ''
        sheet[f"E{row_num}"] = objet.categorie.name if objet.categorie else '' 
        sheet[f"F{row_num}"] = objet.montant

            # Ajouter la photo
        # if objet.photo:  # Assurez-vous que votre modèle Objet a un champ 'photo' avec le chemin de l'image
        #     image_path = objet.photo.path
        #     img = Image(image_path)
        #     img.height = 150  # Ajustez la hauteur de l'image
        #     img.width = 100   # Ajustez la largeur de l'image
        #     cell_coordinates = f"G{row_num}"  # Colonne G pour les photos
        #     sheet.add_image(img, cell_coordinates)
    # Préparer la réponse HTTP
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename="objets.xlsx"'

    workbook.save(response)
    return response

@login_required
def home(request):
    return render(request, 'home.html')

def page_encours(request):
    return render(request, 'Encours_construction.html')

class ProtectedView(LoginRequiredMixin, TemplateView):
    template_name = 'protected.html'
    

class TableauBordView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = Objet
    template_name = 'tableau_bord.html'
    context_object_name = 'objets'
    permission_required = 'biens.view_Objet'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Annoter les objets avec les sous-totaux par rubrique
        context['rubriques'] = Objet.objects.values('rubrique_id').annotate(total_montant=Sum('montant'))
        
        # Ajouter les noms des rubriques
        context['detail_rubriques'] = Rubrique.objects.filter(id__in=[r['rubrique_id'] for r in context['rubriques']])
        
        # Fusionner les sous-totaux et les noms des rubriques et montant_assu
        for rubrique in context['rubriques']:
            rubrique['name'] = next((r.name for r in context['detail_rubriques'] if r.id == rubrique['rubrique_id']), None)
            rubrique['montant_assu'] = next((r.montant_assu for r in context['detail_rubriques'] if r.id == rubrique['rubrique_id']), None)
        
        # Créer le graphique
        categories = [rubrique['name'] for rubrique in context['rubriques']]
        total_montant = [rubrique['total_montant'] for rubrique in context['rubriques']]
        montant_assu = [rubrique['montant_assu'] for rubrique in context['rubriques']]

        x = range(len(categories))

        plt.figure(figsize=(10, 5))
        plt.bar(x, total_montant, width=0.4, label='Montant consommé', color='red', align='center')
        plt.bar([p + 0.4 for p in x], montant_assu, width=0.4, label='Montant d\'assurance', color='green', align='center')
        plt.xlabel('Rubrique Assurance')
        plt.ylabel('Montant')
        plt.title('Répartition contrat assurance')
        plt.xticks([p + 0.2 for p in x], categories)
        plt.legend()

        # Sauvegarder le graphique dans un buffer
        buffer = BytesIO()
        plt.savefig(buffer, format='png')
        buffer.seek(0)

        # Ajouter le graphique au contexte
        # Encoder le graphique en base64
        image_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
        
        context['barchart'] = image_base64
        context['pieces'] = Objet.objects.values('piece').annotate(total_montant=Sum('montant'))
        context['categories'] = Objet.objects.values('categorie__name').annotate(total_montant=Sum('montant'))
        return context

class ObjetListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = Objet
    context_object_name = "objets"
    template_name = 'liste_objets.html'
    permission_required = 'biens.view_Objet'
    paginate_by = 10
    
    def get_queryset(self):
        query = self.request.GET.get('q')
        if query:
            return Objet.objects.filter(
                Q(name__icontains=query) | Q(rubrique__name__icontains=query)
                | Q(categorie__name__icontains=query)
                | Q(piece__icontains=query)
            )
        return Objet.objects.all()
        
class ObjetDetailView(DetailView):
    model = Objet
    template_name = 'detail_objet.html'
    


